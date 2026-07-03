"""Parser dinámico para archivos Excel de rutas Pan Pa Ya.

Estructura del Excel:
- Fila 1: fecha de entrega
- Bloques bidireccionales en columnas A-D (izq) y F-I (der), separadas por col E
- Cada bloque: encabezado de zona → filas de clientes → NOMBRE / PLACA / AUX.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import BinaryIO

import openpyxl

# Columnas 0-indexed
COL_IZQ = (0, 1, 2, 3)  # código, nombre, dirección, notas
COL_DER = (5, 6, 7, 8)
COL_SEP = 4

DRIVER_LABELS = {"NOMBRE", "PLACA", "AUX.", "AUX"}
CODIGO_PATTERN = re.compile(r"^\d{4,7}$")


@dataclass
class PedidoParseado:
    codigo_cliente: str
    nombre_cliente: str
    direccion: str
    notas: str | None = None
    orden: int | None = None


@dataclass
class RutaParseada:
    fecha: date
    nombre_zona: str
    lado: str  # 'izquierda' | 'derecha'
    conductor: str
    placa: str
    auxiliar: str | None = None
    hora_salida: time | None = None
    pedidos: list[PedidoParseado] = field(default_factory=list)


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def _parse_fecha(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value)
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    raise ValueError(f"No se pudo interpretar la fecha: {value!r}")


def _parse_hora(value) -> time | None:
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    text = _cell_str(value)
    if not text:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return None


def _parse_orden(value) -> int | None:
    text = _cell_str(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _is_codigo_cliente(text: str) -> bool:
    return bool(CODIGO_PATTERN.match(text))


def _is_driver_row(label: str) -> bool:
    return label.upper() in DRIVER_LABELS or label.upper().startswith("AUX")


def _is_zone_header(row: list[str], col: int) -> bool:
    """Detecta encabezado de zona en una columna."""
    label = row[col] if col < len(row) else ""
    if not label:
        return False
    if _is_codigo_cliente(label):
        return False
    if _is_driver_row(label):
        return False
    if label.upper() in ("DIRECCION", "DIRECCIÓN", "ORDEN", "MULTIPLAZA"):
        return False
    return True


def _detect_zone(row: list[str], side: str) -> str | None:
    """Detecta nombre de zona en layout estándar (col 0/5) o MULTIAMBIENTE (col 1/6)."""
    if side == "izquierda":
        cols = (0, 1)
        opp_base = 5
    else:
        cols = (5, 6)
        opp_base = 0

    for col in cols:
        if not _is_zone_header(row, col):
            continue
        # Título desplazado (MULTIAMBIENTE): clientes siguen en col base
        if col in (1, 6) and row[cols[0]]:
            continue
        # Si el otro lado tiene código de cliente, es etiqueta interna, no nueva zona
        if _is_codigo_cliente(row[opp_base]):
            continue
        return row[col]
    return None


def _row_starts_new_block(row: list[str]) -> bool:
    return _detect_zone(row, "izquierda") is not None or _detect_zone(row, "derecha") is not None


def _read_row(ws, row_idx: int, max_col: int = 9) -> list[str]:
    return [_cell_str(ws.cell(row=row_idx, column=c + 1).value) for c in range(max_col)]


def _extract_cliente(row: list[str], cols: tuple[int, ...]) -> PedidoParseado | None:
    codigo = row[cols[0]] if cols[0] < len(row) else ""
    if not _is_codigo_cliente(codigo):
        return None
    nombre = row[cols[1]] if cols[1] < len(row) else ""
    direccion = row[cols[2]] if cols[2] < len(row) else ""
    notas = row[cols[3]] if cols[3] < len(row) else ""
    orden_val = _parse_orden(notas)
    notas_final = notas if notas and orden_val is None else None
    return PedidoParseado(
        codigo_cliente=codigo,
        nombre_cliente=nombre,
        direccion=direccion,
        notas=notas_final if notas_final else None,
        orden=orden_val,
    )


def _parse_block(
    rows: list[list[str]],
    col_start: int,
    lado: str,
    fecha: date,
    nombre_zona: str,
) -> RutaParseada | None:
    pedidos: list[PedidoParseado] = []
    conductor = ""
    placa = ""
    auxiliar: str | None = None
    hora_salida: time | None = None

    for row in rows:
        label = row[col_start] if col_start < len(row) else ""

        if _is_driver_row(label):
            valor = row[col_start + 1] if col_start + 1 < len(row) else ""
            hora_celda = row[col_start + 2] if col_start + 2 < len(row) else ""
            upper = label.upper()
            if upper == "NOMBRE":
                conductor = valor
                h = _parse_hora(hora_celda)
                if h:
                    hora_salida = h
            elif upper == "PLACA":
                placa = valor
                h = _parse_hora(hora_celda)
                if h and not hora_salida:
                    hora_salida = h
            elif upper.startswith("AUX"):
                auxiliar = valor or None
                h = _parse_hora(hora_celda)
                if h and not hora_salida:
                    hora_salida = h
            continue

        cliente = _extract_cliente(row, (col_start, col_start + 1, col_start + 2, col_start + 3))
        if cliente:
            pedidos.append(cliente)

    if not conductor and not placa and not pedidos:
        return None

    return RutaParseada(
        fecha=fecha,
        nombre_zona=nombre_zona,
        lado=lado,
        conductor=conductor or "SIN ASIGNAR",
        placa=placa or "SIN PLACA",
        auxiliar=auxiliar,
        hora_salida=hora_salida,
        pedidos=pedidos,
    )


def parse_excel(source: str | Path | BinaryIO) -> list[RutaParseada]:
    """Parsea un archivo Excel y devuelve todas las rutas detectadas."""
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb.active

    fecha = _parse_fecha(ws.cell(row=1, column=1).value)
    rutas: list[RutaParseada] = []
    max_row = ws.max_row

    row_idx = 2
    while row_idx <= max_row:
        row = _read_row(ws, row_idx)

        zona_izq = _detect_zone(row, "izquierda")
        zona_der = _detect_zone(row, "derecha")

        if not zona_izq and not zona_der:
            row_idx += 1
            continue

        block_rows: list[list[str]] = []
        row_idx += 1
        while row_idx <= max_row:
            next_row = _read_row(ws, row_idx)
            if _row_starts_new_block(next_row):
                break
            if all(not c for c in next_row) and row_idx + 1 <= max_row:
                peek = _read_row(ws, row_idx + 1)
                if _row_starts_new_block(peek):
                    break
            block_rows.append(next_row)
            row_idx += 1

        if zona_izq:
            ruta = _parse_block(block_rows, COL_IZQ[0], "izquierda", fecha, zona_izq)
            if ruta:
                rutas.append(ruta)

        if zona_der:
            ruta = _parse_block(block_rows, COL_DER[0], "derecha", fecha, zona_der)
            if ruta:
                rutas.append(ruta)

    return rutas
